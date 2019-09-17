#import modules
import sys
import numpy as np
import poly_fits as pf
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Alignment, Border, Side
import json

def display_coef(var1_name, var2_name, func_name, n_vec, a, r2, rms, rmsn, row, column, sheet):
	
	last_row = row + 2 + n_vec[1]
	if last_row < row + 7: last_row = row + 7
	last_column = column + 4 + n_vec[0]
	
	for i in range(column,last_column+1):
		sheet.column_dimensions[get_column_letter(i)].width = 11.
	al = Alignment(horizontal = 'center', vertical = 'center')
	
	sheet.merge_cells(start_row = row, start_column = column, end_row = row + 1, end_column = column + 1)
	c = sheet.cell( row = row, column = column)
	c.value = func_name
	c.alignment = al
	
	sheet.merge_cells(start_row = row, start_column = column + 2, end_row = row, end_column = column + 2 + n_vec[0])
	c = sheet.cell( row = row, column = column + 2)
	c.value = var1_name
	c.alignment = al
	
	sheet.merge_cells(start_row = row + 2, start_column = column, end_row = row + 2 + n_vec[1], end_column = column)
	c = sheet.cell( row = row + 2, column = column)
	c.value = var2_name
	c.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
	
	#create constant list
	data_coef_str = []
	for i in range(100):
		data_coef_str.append( 'C{}'.format(i) )
	
	for i in range(n_vec[0]+1):
		c = sheet.cell( row = row + 1, column = column + 2 + i)
		c.value = data_coef_str[i]
		c.alignment = al
	for i in range(n_vec[1]+1):
		c = sheet.cell( row = row + 2 + i, column = column + 1)
		c.value = data_coef_str[i]
		c.alignment = al
	
	for n in range(n_vec[0]+1):
		for m in range(n_vec[1]+1):
			j = pf.compose_j([n,m], n_vec)
			if a[j] != 0.:
				c = sheet.cell( row = row + 2 + m, column = column + 2 + n)
				c.value = a[j]
				c.alignment = al
	
	c = sheet.cell( row = row, column = column + 2 + n_vec[0] + 2)
	c.value = 'R^2'
	c.alignment = al
	c = sheet.cell( row = row + 1, column = column + 2 + n_vec[0] + 2)
	c.value = r2
	c.alignment = al
	
	c = sheet.cell( row = row + 3, column = column + 2 + n_vec[0] + 2)
	c.value = 'RMS Error'
	c.alignment = al
	c = sheet.cell( row = row + 4, column = column + 2 + n_vec[0] + 2)
	c.value = rms
	c.alignment = al
	
	c = sheet.cell( row = row + 6, column = column + 2 + n_vec[0] + 2)
	c.value = 'RMSN Error'
	c.alignment = al
	c = sheet.cell( row = row + 7, column = column + 2 + n_vec[0] + 2)
	c.value = rmsn
	c.alignment = al
	
	ran = get_column_letter(column)+str(row)+':'+get_column_letter(last_column) + str(last_row)
	rows = sheet[ran]
	border = Border(left = Side(border_style = 'thick'),
	                right = Side(border_style = 'thick'),
	                top = Side(border_style = 'thick'),
	                bottom = Side(border_style = 'thick'))
	bottom = Border(bottom = border.bottom)
	right = Border(right = border.right)
	left = Border(left = border.left)
	top = Border(top = border.top)
	for c in rows[0]:
		c.border = c.border + top
	for c in rows[-1]:
		c.border = c.border + bottom
	for ro in rows:
		l = ro[0]
		r = ro[-1]
		l.border = l.border + left
		r.border += right

#input airfoil file
if len(sys.argv) > 1:
	input_filename = sys.argv[1]
else:
	input_filename = input('\n'*2+'Enter name of file for input: ')
ff = open(input_filename, 'r')
f = json.load(ff)
ff.close()

print('Reading in raw data')

# load excel file
filename_excel = f['step 1']['excel file']
wb = load_workbook(filename_excel+'.xlsx')
ws_coef = wb.active
ws_coef.title = 'Function Coefficients'
ws_coef_row = 1

#setup flap deflection range
df_min = f['step 1']['flap deflection']['min']
df_max = f['step 1']['flap deflection']['max']
delta_df = f['step 1']['flap deflection']['delta']
N_df = int((df_max-df_min)/delta_df) + 1

#setup aoa range
aoa_min = f['step 1']['angle of attack']['min']
aoa_max = f['step 1']['angle of attack']['max']
delta_aoa = f['step 1']['angle of attack']['delta']
N_aoa = int((aoa_max-aoa_min)/delta_aoa) + 1

#poly_fit variables
x = []
CL = []
CD = []
Cm = []
Ch = []
n_vec_CL = f['step 2']['CL']
n_vec_CD = f['step 2']['CD']
n_vec_Cm = f['step 2']['Cm']
n_vec_Ch = f['step 2']['Ch']

#read xfoil setup variables
ncrit = f['step 1']['ncrit']
xtr = f['step 1']['xtr']
NACA = f['step 1']['NACA']
grid = f['step 1']['grid']
Re = f['step 1']['Re']
xhinge = f['step 1']['xf']
Niter = f['step 1']['Niter']

#write xfoil setup variables to excel file
for i in range(2,15):
	ws_coef.column_dimensions[get_column_letter(i)].width = 11.

al = Alignment(horizontal = 'center', vertical = 'center')
ws_coef.merge_cells( start_row = 2, start_column = 2, end_row = 2, end_column = 14 )
c = ws_coef.cell( row = 2, column = 2)
c.value = 'XFOIL settings'
c.alignment = al

c = ws_coef.cell( row = 3, column = 2)
c.value = 'Re'
c.alignment = al

c = ws_coef.cell( row = 3, column = 3)
c.value = 'x, trip'
c.alignment = al

c = ws_coef.cell( row = 3, column = 4)
c.value = 'Ncrit'
c.alignment = al

c = ws_coef.cell( row = 3, column = 5)
c.value = 'airfoil grid'
c.alignment = al

c = ws_coef.cell( row = 3, column = 6)
c.value = 'NACA'
c.alignment = al

c = ws_coef.cell( row = 3, column = 7)
c.value = 'x, hinge'
c.alignment = al

c = ws_coef.cell( row = 3, column = 8)
c.value = 'iterations'
c.alignment = al

c = ws_coef.cell( row = 3, column = 9)
c.value = 'flap, min'
c.alignment = al

c = ws_coef.cell( row = 3, column = 10)
c.value = 'flap, max'
c.alignment = al

c = ws_coef.cell( row = 3, column = 11)
c.value = 'flap, step'
c.alignment = al

c = ws_coef.cell( row = 3, column = 12)
c.value = 'aoa, min'
c.alignment = al

c = ws_coef.cell( row = 3, column = 13)
c.value = 'aoa, max'
c.alignment = al

c = ws_coef.cell( row = 3, column = 14)
c.value = 'aoa, step'
c.alignment = al

c = ws_coef.cell( row = 4, column = 2)
c.value = Re
c.alignment = al

c = ws_coef.cell( row = 4, column = 3)
c.value = xtr
c.alignment = al

c = ws_coef.cell( row = 4, column = 4)
c.value = ncrit
c.alignment = al

c = ws_coef.cell( row = 4, column = 5)
c.value = grid
c.alignment = al

c = ws_coef.cell( row = 4, column = 6)
c.value = NACA
c.alignment = al

c = ws_coef.cell( row = 4, column = 7)
c.value = xhinge
c.alignment = al

c = ws_coef.cell( row = 4, column = 8)
c.value = Niter
c.alignment = al

c = ws_coef.cell( row = 4, column = 9)
c.value = df_min
c.alignment = al

c = ws_coef.cell( row = 4, column = 10)
c.value = df_max
c.alignment = al

c = ws_coef.cell( row = 4, column = 11)
c.value = delta_df
c.alignment = al

c = ws_coef.cell( row = 4, column = 12)
c.value = aoa_min
c.alignment = al

c = ws_coef.cell( row = 4, column = 13)
c.value = aoa_max
c.alignment = al

c = ws_coef.cell( row = 4, column = 14)
c.value = delta_aoa
c.alignment = al

ran = 'B2:N4'
rows = ws_coef[ran]
border = Border(left = Side(border_style = 'thick'),
                right = Side(border_style = 'thick'),
                top = Side(border_style = 'thick'),
                bottom = Side(border_style = 'thick'))
bottom = Border(bottom = border.bottom)
right = Border(right = border.right)
left = Border(left = border.left)
top = Border(top = border.top)
for c in rows[0]:
	c.border = c.border + top
for c in rows[-1]:
	c.border = c.border + bottom
for ro in rows:
	l = ro[0]
	r = ro[-1]
	l.border = l.border + left
	r.border += right

count = 0

#loop through the flap deflection range
for i in range(N_df):
	#set the flap deflection for this loop iteration
	df = df_min + float(i) * delta_df
	
	#determine the filename
	if df < 0.:
		sgn = 'n'
	else:
		sgn = 'p'
	filename = '{}{:0>2d}'.format(sgn,int(abs(df)))       # filename cannot be greater than 6 characters long
	
	#create a new excel sheet
	ws = wb[filename]
	ws_row = 1
	
	#initialize some variable to extract data from the xfoil run
	N = ws.max_row - 1
	#extract data from xfoil run and put it on the excel sheet
	for j in range(N):
		#extract data from sheet
		x.append([])
		ws_row += 1
		col = 1
		x[count].append( ws.cell(row = ws_row, column = col).value )
		x[count].append( df )
		col += 1
		# aoa_deg[j] = ws.cell(row = ws_row, column = col).value #, value = aoa_deg[j])
		col += 1
		CL.append( ws.cell(row = ws_row, column = col).value )
		col += 1
		CD.append( ws.cell(row = ws_row, column = col).value )
		col += 1
		Cm.append( ws.cell(row = ws_row, column = col).value )
		col += 1
		Ch.append( ws.cell(row = ws_row, column = col).value )
		
		count += 1

print('Performing CL curve fit')
aCL, rCL = pf.multidimensional_poly_fit(n_vec_CL, x, CL, sym_same = [[0,1]])
rmsCL, rmsnCL = pf.multidimensional_rms(x, CL, aCL, n_vec_CL)
print('Performing CD curve fit')
aCD, rCD = pf.multidimensional_poly_fit(n_vec_CD, x, CD, sym_diff = [[0,1]])
rmsCD, rmsnCD = pf.multidimensional_rms(x, CD, aCD, n_vec_CD)

#############################################################################
# print('generating 3D plot')
# import matplotlib.pyplot as plt
# from mpl_toolkits.mplot3d import Axes3D

# aoa_plot = []
# df_plot = []
# CD_plot = []
# f_plot = []
# #error_plot = []
# for i,xx in enumerate(x):
	# if i%30 == 0:
		# aoa_plot.append( xx[0] * 180. / np.pi )
		# df_plot.append( xx[1] )
		# CD_plot.append( CD[i] )

# aoa_plot1 = []
# df_plot1 = []
# for i in np.linspace(-15.,15.,10) * np.pi / 180.:
	# for j in np.linspace(-20.,20.,100):
		# xx = [i,j]
		# aoa_plot1.append( i * 180. / np.pi )
		# df_plot1.append( j )
		# f_plot.append( pf.multidimensional_poly_func(aCD,n_vec_CD,xx) )
		# #error_plot.append( CD[i] - pf.multidimensional_poly_func(aCD,n_vec_CD,xx) )

# fig1 = plt.figure()
# ax1 = fig1.add_subplot(111, projection='3d')
# ax1.scatter(aoa_plot,df_plot,CD_plot,'b')
# ax1.scatter(aoa_plot1,df_plot1,f_plot,'r')
# ax1.set_xlabel('aoa')
# ax1.set_ylabel('df')
# ax1.set_zlabel('CD')
# #ax1.set_title('Symmetric Inner')
# #ax1.set_xlim3d(-.05,.05)

# #fig2 = plt.figure()
# #ax2 = fig2.add_subplot(111, projection = '3d')
# #ax2.scatter(aoa_plot, df_plot, error_plot, 'r')
# #ax2.set_xlabel('aoa')
# #ax2.set_ylabel('df')
# #ax2.set_zlabel('error')

# plt.show()


#############################################################################

print('Performing Cm curve fit')
aCm, rCm = pf.multidimensional_poly_fit(n_vec_Cm, x, Cm)
rmsCm, rmsnCm = pf.multidimensional_rms(x, Cm, aCm, n_vec_Cm)
print('Performing Ch curve fit')
aCh, rCh = pf.multidimensional_poly_fit(n_vec_Ch, x, Ch)
rmsCh, rmsnCh = pf.multidimensional_rms(x, Ch, aCh, n_vec_Ch)





row = 6
column = 2
display_coef('aoa', 'control deflection', 'CL', n_vec_CL, aCL, rCL, rmsCL, rmsnCL, row, column, ws_coef)
display_coef('aoa', 'control deflection', 'CD', n_vec_CD, aCD, rCD, rmsCD, rmsnCD, row + n_vec_CL[1] + 4, column, ws_coef)
display_coef('aoa', 'control deflection', 'Cm', n_vec_Cm, aCm, rCm, rmsCm, rmsnCm, row + n_vec_CL[1] + 4 + n_vec_CD[1] + 4, column, ws_coef)
display_coef('aoa', 'control deflection', 'Ch', n_vec_Ch, aCh, rCh, rmsCh, rmsnCh, row + n_vec_CL[1] + 4 + n_vec_CD[1] + 4 + n_vec_Cm[1] + 4, column, ws_coef)

print('writing results to excel file and generating airfoil json file')

#loop through the flap deflection range
for i in range(N_df):
	#set the flap deflection for this loop iteration
	df = df_min + float(i) * delta_df
	
	#determine the filename
	if df < 0.:
		sgn = 'n'
	else:
		sgn = 'p'
	filename = '{}{:0>2d}'.format(sgn,int(abs(df)))       # filename cannot be greater than 6 characters long
	
	#create a new excel sheet
	ws = wb[filename]
	ws_row = 1
	
	#initialize some variable to extract data from the xfoil run
	N = ws.max_row - 1
	aoa_raw_deg = [None] * N
	aoa_raw_rad = np.zeros((N,2))
	CL_raw = [None] * N
	CD_raw = [None] * N
	Cm_raw = [None] * N
	Ch_raw = [None] * N
	#extract data from xfoil run and put it on the excel sheet
	for j in range(N):
		#extract data from sheet
		ws_row += 1
		col = 1
		aoa_raw_rad[j,0] = ws.cell(row = ws_row, column = col).value
		aoa_raw_rad[j,1] = df
		col += 1
		aoa_raw_deg[j] = ws.cell(row = ws_row, column = col).value
		col += 1
		CL_raw[j] = ws.cell(row = ws_row, column = col).value
		col += 1
		CD_raw[j] = ws.cell(row = ws_row, column = col).value
		col += 1
		Cm_raw[j] = ws.cell(row = ws_row, column = col).value
		col += 1
		Ch_raw[j] = ws.cell(row = ws_row, column = col).value
	
	#create trendline data
	N_plot = 100
	CD_anal = [None] * N_plot
	CL_anal = [None] * N_plot
	Cm_anal = [None] * N_plot
	Ch_anal = [None] * N_plot
	aoa_anal_deg = np.linspace(aoa_min,aoa_max,N_plot)
	aoa_anal_rad = aoa_anal_deg * np.pi / 180.
	
	r2CL = pf.multidimensional_r2(aCL, n_vec_CL, aoa_raw_rad, CL_raw)
	r2CD = pf.multidimensional_r2(aCD, n_vec_CD, aoa_raw_rad, CD_raw)
	r2Cm = pf.multidimensional_r2(aCm, n_vec_Cm, aoa_raw_rad, Cm_raw)
	r2Ch = pf.multidimensional_r2(aCh, n_vec_Ch, aoa_raw_rad, Ch_raw)
	
	#add trendline date to the excel sheet
	row = 1
	col = 8
	ws.cell(row = row, column = col, value = 'aoa (rad)')
	col += 1
	ws.cell(row = row, column = col, value = 'aoa (deg)')
	col += 1
	ws.cell(row = row, column = col, value = 'CL(aoa,df) r2 = {:.5f}'.format(r2CL))
	col += 1
	ws.cell(row = row, column = col, value = 'CD(aoa,df) r2 = {:.5f}'.format(r2CD))
	col += 1
	ws.cell(row = row, column = col, value = 'Cm(aoa,df) r2 = {:.5f}'.format(r2Cm))
	col += 1
	ws.cell(row = row, column = col, value = 'Ch(aoa,df) r2 = {:.5f}'.format(r2Ch))
	for row in range(N_plot):
		col = 8
		ws.cell(row = row+2, column = col, value = aoa_anal_rad[row])
		col += 1
		ws.cell(row = row+2, column = col, value = aoa_anal_deg[row])
		col += 1
		ws.cell(row = row+2, column = col, value = pf.multidimensional_poly_func(aCL, n_vec_CL, [aoa_anal_rad[row], df]))
		col += 1
		ws.cell(row = row+2, column = col, value = pf.multidimensional_poly_func(aCD, n_vec_CD, [aoa_anal_rad[row], df]))
		col += 1
		ws.cell(row = row+2, column = col, value = pf.multidimensional_poly_func(aCm, n_vec_Cm, [aoa_anal_rad[row], df]))
		col += 1
		ws.cell(row = row+2, column = col, value = pf.multidimensional_poly_func(aCh, n_vec_Ch, [aoa_anal_rad[row], df]))
	
	#set up charts
	point = 12700
	#set up CL chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 9).value
	chart.y_axis.title = 'CL'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws.max_row)
	yvalues = Reference(ws, min_col = 3, min_row = 1, max_row = ws.max_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#CL(aoa) series
	xvalues = Reference(ws, min_col = 9, min_row = 2, max_row = N_plot+1)
	yvalues = Reference(ws, min_col =10, min_row = 1, max_row = N_plot+1)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.graphicalProperties.line.solidFill = 'FF0000'
	series.graphicalProperties.line.width = point * 2
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "B2")
	
	#set up CD chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 9).value
	chart.y_axis.title = 'CD'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws.max_row)
	yvalues = Reference(ws, min_col = 4, min_row = 1, max_row = ws.max_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#CD(aoa) series
	xvalues = Reference(ws, min_col =  9, min_row = 2, max_row = N_plot+1)
	yvalues = Reference(ws, min_col = 11, min_row = 1, max_row = N_plot+1)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.graphicalProperties.line.solidFill = 'FF0000'
	series.graphicalProperties.line.width = point * 2
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "L2")
	
	#set up Cm chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 9).value
	chart.y_axis.title = 'Cm'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws.max_row)
	yvalues = Reference(ws, min_col = 5, min_row = 1, max_row = ws.max_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#Cm(aoa) series
	xvalues = Reference(ws, min_col =  9, min_row = 2, max_row = N_plot+1)
	yvalues = Reference(ws, min_col = 12, min_row = 1, max_row = N_plot+1)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.graphicalProperties.line.solidFill = 'FF0000'
	series.graphicalProperties.line.width = point * 2
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "V2")
	
	#set up Ch chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 9).value
	chart.y_axis.title = 'Chinge'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws.max_row)
	yvalues = Reference(ws, min_col = 6, min_row = 1, max_row = ws.max_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#Ch(aoa) series
	xvalues = Reference(ws, min_col =  9, min_row = 2, max_row = N_plot+1)
	yvalues = Reference(ws, min_col = 13, min_row = 1, max_row = N_plot+1)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.graphicalProperties.line.solidFill = 'FF0000'
	series.graphicalProperties.line.width = point * 2
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "B31")

wb.save(filename_excel+'_final'+'.xlsx')

#prep for airfoil database json file
filename = filename_excel
#add common data
json_data = { filename : ''}
json_data[filename] = { 'properties' : ''}
data = {}
data['type'] = 'polynomial'
data['is_function'] = 1
#create constant list
data_coef_str = []
for i in range(100):
	data_coef_str.append( 'C{}'.format(i) )
#setup coefficients for CL
data['CL'] = {}
for n in range(n_vec_CL[0]+1):
	data['CL'][data_coef_str[n]] = {}
	for m in range(n_vec_CL[1]+1):
		j = pf.compose_j([n,m], n_vec_CL)
		if aCL[j] != 0.: data['CL'][data_coef_str[n]][data_coef_str[m]] = aCL[j]
	if data['CL'][data_coef_str[n]] == {}: data['CL'][data_coef_str[n]][data_coef_str[0]] = 0.
#setup coefficients for CD
data['CD'] = {}
for n in range(n_vec_CD[0]+1):
	data['CD'][data_coef_str[n]] = {}
	for m in range(n_vec_CD[1]+1):
		j = pf.compose_j([n,m], n_vec_CD)
		if aCD[j] != 0.: data['CD'][data_coef_str[n]][data_coef_str[m]] = aCD[j]
	if data['CD'][data_coef_str[n]] == {}: data['CD'][data_coef_str[n]][data_coef_str[0]] = 0.
#setup coefficients for Cm
data['Cm'] = {}
for n in range(n_vec_Cm[0]+1):
	data['Cm'][data_coef_str[n]] = {}
	for m in range(n_vec_Cm[1]+1):
		j = pf.compose_j([n,m], n_vec_Cm)
		if aCm[j] != 0.: data['Cm'][data_coef_str[n]][data_coef_str[m]] = aCm[j]
	if data['Cm'][data_coef_str[n]] == {}: data['Cm'][data_coef_str[n]][data_coef_str[0]] = 0.
#setup coefficients for Ch
data['Ch'] = {}
for n in range(n_vec_Ch[0]+1):
	data['Ch'][data_coef_str[n]] = {}
	for m in range(n_vec_Ch[1]+1):
		j = pf.compose_j([n,m], n_vec_Ch)
		if aCh[j] != 0.: data['Ch'][data_coef_str[n]][data_coef_str[m]] = aCh[j]
	if data['Ch'][data_coef_str[n]] == {}: data['Ch'][data_coef_str[n]][data_coef_str[0]] = 0.
data['CL_max'] = ''
data['Comments'] = 'All angles in radians and slopes in 1/radians'
json_data[filename]['properties'] = data
with open(filename+'.json', 'w') as data_file:
	json.dump(json_data, data_file, indent = '\t')


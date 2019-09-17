#import modules
import sys
import VortexPanel as VP
import airfoil as af
import xfoil as xfoil
import numpy as np
# import airfoil_database_mod as zach
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import json

#setup excel file
wb = Workbook()

#input airfoil file
if len(sys.argv) > 1:
	input_filename = sys.argv[1]
else:
	input_filename = input('\n'*2+'Enter name of file for input: ')
ff = open(input_filename, 'r')
f = json.load(ff)
ff.close()

#setup airfoil constants
m = VP.driver()
NACA = f['step 1']['NACA']
xf = f['step 1']['xf']

#setup flap deflection range
df_min = f['step 1']['flap deflection']['min']
df_max = f['step 1']['flap deflection']['max']
delta_df = f['step 1']['flap deflection']['delta']
N_df = int((df_max-df_min)/delta_df) + 1

#setup aoa range
aoa_min = f['step 1']['angle of attack']['min']
aoa_max = f['step 1']['angle of attack']['max']
delta_aoa = f['step 1']['angle of attack']['delta']

#setup ncrit
ncrit = f['step 1']['ncrit']

#setup xtr range
xtr = f['step 1']['xtr']

#setup the Reynolds number
Re = f['step 1']['Re']

#set grid size
grid = f['step 1']['grid']

#set number of iterations until not converged
niter = f['step 1']['Niter']

#loop through the flap deflection range
for i in range(N_df):
	#set the flap deflection for this loop iteration
	df = df_min + float(i) * delta_df
	
	#determine the filename
	if df < 0.:
		sgn = 'n'
	else:
		sgn = 'p'
	filename = '{}{:0>2d}'.format(sgn,int(abs(df)))       # filename cannot be greater than 6 characters long
	
	#create the airfoil and run xfoil
	myaf = af.airfoil(naca=NACA,xf=xf,flapType=0,delta=df)
	m.add_airfoil(myaf)
	m.write_xfoil(filename = filename+'.txt')
	deflection_set = xfoil.generate_machup_from_xfoil(airfoils			= [filename+'.txt'],
									alpha_min			= aoa_min,
									alpha_max			= aoa_max,
									delta_alpha			= delta_aoa,
									reynolds_number		= Re,
									grid_size			= grid,
									niter				= niter,
									ncrit				= ncrit,
									xtrt				= xtr,
									xtrb				= xtr,
									flap				= df,
									xf					= xf,
									yf					= 0.0)
	
	#initialize some variable to extract data from the xfoil run
	N = len(deflection_set)
	aoa_deg = [None] * N
	aoa_rad = [None] * N
	CL = [None] * N
	CD = [None] * N
	Cm = [None] * N
	Ch = [None] * N
	# cl2, cl3, cl4, aoa2, aoa3, aoa4, cdcl, cdcl2, claoa, claoa2, cmaoa, cmaoa2 = [0.] * 12
	
	#create a new excel sheet
	ws = wb.create_sheet(title = filename)
	ws_row = 1
	col = 1
	ws.cell(row = ws_row, column = col, value = 'aoa (rad)')
	col += 1
	ws.cell(row = ws_row, column = col, value = 'aoa (deg)')
	col += 1
	ws.cell(row = ws_row, column = col, value = 'CL')
	col += 1
	ws.cell(row = ws_row, column = col, value = 'CD')
	col += 1
	ws.cell(row = ws_row, column = col, value = 'Cm')
	col += 1
	ws.cell(row = ws_row, column = col, value = 'CHinge')
	
	#extract data from xfoil run and put it on the excel sheet
	for j in range(N):
		#extract data from xfoil
		aoa_deg[j],CL[j],CD[j],Cm[j],Ch[j] = deflection_set[j]
		aoa_rad[j] = aoa_deg[j] * np.pi / 180.
		#add data to the excel sheet
		ws_row += 1
		col = 1
		ws.cell(row = ws_row, column = col, value = aoa_rad[j])
		col += 1
		ws.cell(row = ws_row, column = col, value = aoa_deg[j])
		col += 1
		ws.cell(row = ws_row, column = col, value = CL[j])
		col += 1
		ws.cell(row = ws_row, column = col, value = CD[j])
		col += 1
		ws.cell(row = ws_row, column = col, value = Cm[j])
		col += 1
		ws.cell(row = ws_row, column = col, value = Ch[j])
	
	#set up charts
	point = 12700
	#set up CL chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 2).value
	chart.y_axis.title = 'CL'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws_row)
	yvalues = Reference(ws, min_col = 3, min_row = 1, max_row = ws_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "B2")
	
	#set up CD chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 2).value
	chart.y_axis.title = 'CD'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws_row)
	yvalues = Reference(ws, min_col = 4, min_row = 1, max_row = ws_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "L2")
	
	#set up Cm chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 2).value
	chart.y_axis.title = 'Cm'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws_row)
	yvalues = Reference(ws, min_col = 5, min_row = 1, max_row = ws_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "V2")
	
	#set up Ch chart
	chart = ScatterChart()
	chart.x_axis.title = ws.cell(row = 1, column = 2).value
	chart.y_axis.title = 'Chinge'
	chart.x_axis.scaling.min = aoa_min
	chart.x_axis.scaling.max = aoa_max
	chart.y_axis.crosses = "min"
	chart.x_axis.crosses = "min"
	chart.height = 15.
	#raw data series
	xvalues = Reference(ws, min_col = 2, min_row = 2, max_row = ws_row)
	yvalues = Reference(ws, min_col = 6, min_row = 1, max_row = ws_row)
	series = Series(yvalues, xvalues, title_from_data = True)
	series.marker.symbol = 'circle'
	series.marker.size = 5.
	series.graphicalProperties.line.noFill = True
	series.marker.graphicalProperties.noFill = True
	series.marker.graphicalProperties.line.solidFill = '000000'
	series.marker.graphicalProperties.line.width = point
	chart.series.append(series)
	#add chart to sheet
	ws.add_chart(chart, "B31")

#prompt user for filename and save results
wb.save(f['step 1']['excel file']+'.xlsx')

